import os
import re
import threading
from datetime import datetime

import fitz
import openpyxl.utils
from flask import Flask
from flask import make_response
from flask import request
from flask import send_file
from flask_cors import CORS
from flask_cors import cross_origin
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

app = Flask(__name__)
CORS(app)

US_FORMAT = [",", "."]
EU_FORMAT = [".", ","]
OUTPUT_FOLDER = './output/'
output_filepath = os.path.join(OUTPUT_FOLDER, "output.xlsx")
ocr = False


def read_pdf_lines(filename):
    doc = fitz.open(filename)
    for page in doc:
        lines = page.get_text("text").split("\n")
        for line in lines:
            yield line.strip()


def replace_consecutive_pipes(text):
    pattern = r"\|{4,}"  # Matches 4 or more consecutive '|'
    replacement = "\n"
    return re.sub(pattern, replacement, text)


def transform_floats(sheet: Worksheet, column: int, format: list = US_FORMAT):
    column_letter = openpyxl.utils.get_column_letter(column)

    if len(column_letter) > 1:
        raise ValueError("the column name is wrong!")

    column_range = sheet[f"{column_letter}2:{column_letter}" + str(sheet.max_row)]

    transformed_values = [
        format_numbers(cell[0].value, format) if cell[0].value else None
        for cell in column_range
    ]

    # Write the transformed values back to the column
    for i, transformed_value in enumerate(transformed_values):
        sheet.cell(row=i + 2, column=column, value=transformed_value)


def transform_dates(sheet: Worksheet, column: int, desired_format: str = "%d/%m/%Y"):
    # Specify the column to convert (e.g., column B)
    column_letter = openpyxl.utils.get_column_letter(column)

    # Iterate over the cells in the column (excluding the title)
    for cell in sheet[column_letter][1 : sheet.max_row]:
        # Convert the cell value to a date
        try:
            cell.value = datetime.strptime(cell.value, "%d/%m/%Y").strftime(
                desired_format
            )
        except ValueError:
            # Handle the error gracefully
            print("couldn't transform this date")
            # Take appropriate action for an invalid date


def transform_ints(sheet: Worksheet, column: int):
    # Specify the column to convert (e.g., column B)
    column_letter = openpyxl.utils.get_column_letter(column)

    # Iterate over the cells in the column (excluding the title)
    for cell in sheet[column_letter][1 : sheet.max_row]:
        # Convert the cell value to a date
        cell.value = int(cell.value)


def format_numbers(nombre, format: list):
    return float(nombre.replace(format[0], "").replace(format[1], format[0]))


def sum_column(sheet: Worksheet, column: int):
    # Calculate the sum of the column
    total = 0
    column_letter = openpyxl.utils.get_column_letter(column)
    for cell in sheet[column_letter][1:]:
        if cell.value is not None:
            total += cell.value

    # Get the letter of the last row in the specified column
    last_row = sheet.max_row + 1

    # Write the sum to the cell under the last row
    sum_cell = sheet[column_letter + str(last_row)]
    sum_cell.value = total


def extract_data(file_path: str, show: bool = False):
    global ocr
    ocr = False
    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Select the active sheet
    sheet = workbook.active

    # Add data to the sheet
    sheet["A1"] = "NUMERO DE COMMANDE"
    sheet["B1"] = "DATE"
    sheet["C1"] = "NUMERO BO"
    sheet["D1"] = "REFERENCE"
    sheet["E1"] = "DESIGNATION"
    sheet["F1"] = "CODE"
    sheet["G1"] = "POIDS NET"
    sheet["H1"] = "QUANTITE"
    sheet["I1"] = "PRIX UNITAIRE"
    sheet["J1"] = "LR"
    sheet["K1"] = "MONTANT"
    sheet["L1"] = "CODE"

    org_lines = []
    line_ = ""

    for line in read_pdf_lines(file_path):
        if line.startswith("|"):
            line_ += "|".join(
                [split_.strip() for split_ in line.replace("\n|", "|").split("|")]
            )
        else:
            line_ += "\n" + line

    for line in replace_consecutive_pipes(line_).split("\n"):
        if line.count("|") == 11:
            org_lines.append(line)

    if show:
        print(org_lines)

    if len(org_lines) < 1:
        ocr = True
        return

    for index, line in enumerate(org_lines):
        for column_number, partie in enumerate(line.split("|")):
            sheet[f"{openpyxl.utils.get_column_letter(column_number+1)}{index + 2}"] = (
                partie
            )

    transform_dates(sheet, 2)

    for index in [7, 9, 11, 8]:
        transform_floats(sheet, index, EU_FORMAT)

    sum_column(sheet, 11)

    to_delete = [12, 10, 7, 6]
    to_delete.sort(reverse=True)
    for index in to_delete:
        sheet.delete_cols(index)

    for column in sheet.columns:
        # Get the column letter
        column_letter = column[0].column_letter

        # Double the current column width
        sheet.column_dimensions[column_letter].width = (
            sheet.column_dimensions[column_letter].width * 1.5
        )

    alignment = Alignment(horizontal="center", vertical="center")
    font = Font(bold=True)
    # Iterate over the cells in the row
    for cell in sheet[1]:
        # Apply bold font style
        cell.font = font
        cell.alignment = alignment

    workbook.save(output_filepath)


@app.route("/api/convert", methods=["POST"])
@cross_origin()
def convert():
    pdf_file = request.files["pdf"]
    pdf_file.save(pdf_file.filename)
    pdf_file = pdf_file.filename

    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)

    if os.path.exists(output_filepath):
        os.remove(output_filepath)

    thread = threading.Thread(target=extract_data, args=(pdf_file, False))
    thread.start()
    thread.join()

    if os.path.exists(pdf_file):
        os.remove(pdf_file)

    if not ocr:
        return send_file(
            output_filepath,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        return make_response("")


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=8080)
